import openpyxl
import pytest

from scripts.preprocess.lavidaagranel import (
    Config,
    MappedResult,
    RawRow,
    iter_rows,
    load_config,
    map_row,
    normalize,
)


def _build_fixture_xlsx(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pedidos Grupo Consumo"
    ws.append(["LA VIDA A GRANEL"])
    ws.append(["PEDIDOS GRUPO DE CONSUMO"])
    ws.append(["Fecha"])
    ws.append([])
    ws.append(["PRODUCTO", "PRECIO", "UNIDAD",
               "FAMILIA 1", "FAMILIA 2", "FAMILIA 3", "FAMILIA 4", "FAMILIA 5",
               "FAMILIA 6", "FAMILIA 7", "FAMILIA 8", "FAMILIA 9", "FAMILIA 10",
               "TOTAL"])
    ws.append(["📁 ALGAS"])
    ws.append(["ALGA KOMBU ECO 25G", 2.5, "Unidades"])
    ws.append(["ALGA NORI 25G", 2.95, "Unidades"])
    ws.append(["📁 LEGUMBRES"])
    ws.append(["GARBANZO ECO", 3.20, "kg"])
    ws.append(["📁 MYSTERY"])
    ws.append(["UNKNOWN ITEM", 1.00, "kg"])
    ws.append([None, None, None])
    ws.append(["📁 ALGAS"])
    ws.append(["BROKEN PRICE", "", "kg"])
    ws.append(["BROKEN UNIT", 1.00, ""])
    wb.save(path)


def test_normalize_basic_title_case():
    assert normalize("ALGA KOMBU ECO") == "Alga Kombu Eco"


def test_normalize_digit_letter_suffix_lowercased():
    assert normalize("ALGA KOMBU ECO 25G") == "Alga Kombu Eco 25g"
    assert normalize("VEER 33CL") == "Veer 33cl"
    assert normalize("AGUA 1L") == "Agua 1l"


def test_normalize_collapses_whitespace():
    assert normalize("  ALGA   KOMBU  ") == "Alga Kombu"


def test_normalize_preserves_accents():
    assert normalize("ALIÑO ESPAÑOL") == "Aliño Español"


def test_normalize_does_not_touch_pure_digit_token():
    assert normalize("PACK 12 UNIDADES") == "Pack 12 Unidades"


def test_normalize_does_not_touch_pure_letter_token():
    assert normalize("CAFE MOLIDO") == "Cafe Molido"


def test_load_config_minimal(fixtures_dir):
    cfg = load_config(fixtures_dir / "lavidaagranel_min.yaml")
    assert isinstance(cfg, Config)
    assert cfg.productor == "La Vida a Granel"
    assert cfg.productor_id == ""
    assert cfg.categorias["📁 ALGAS"] == "Algas y plantas acuáticas"
    assert cfg.unidades["kg"] == {"granel": True, "pesar": True, "descripcion": ""}
    assert cfg.unidades["unidades"]["granel"] is False
    assert cfg.defaults == {
        "destacado": False,
        "temporada": False,
        "precio_final": "",
        "precio_productor": "",
    }


def test_load_config_rejects_unknown_top_level_key(tmp_path):
    bad = tmp_path / "bad.yaml"
    bad.write_text(
        "productor: X\n"
        "categorias: {}\n"
        "unidades: {}\n"
        "defaults: {destacado: false, temporada: false, precio_final: '', precio_productor: ''}\n"
        "surprise: 1\n",
        encoding="utf-8",
    )
    with pytest.raises(ValueError, match="unknown top-level keys"):
        load_config(bad)


def test_load_config_skips_null_category_values(tmp_path):
    p = tmp_path / "c.yaml"
    p.write_text(
        "productor: X\n"
        "categorias:\n"
        "  '📁 ALGAS': { name: 'Algas y plantas acuáticas' }\n"
        "  '📁 MYSTERY': null\n"
        "unidades: { kg: { granel: true, pesar: true, descripcion: '' } }\n"
        "defaults: { destacado: false, temporada: false, precio_final: '', precio_productor: '' }\n",
        encoding="utf-8",
    )
    cfg = load_config(p)
    assert "📁 MYSTERY" not in cfg.categorias
    assert cfg.categorias["📁 ALGAS"] == "Algas y plantas acuáticas"


def test_iter_rows_yields_section_aware_rows(tmp_path):
    xlsx = tmp_path / "min.xlsx"
    _build_fixture_xlsx(xlsx)
    rows = list(iter_rows(xlsx))
    productos = [r for r in rows if r.kind == "product"]
    assert [r.producto for r in productos] == [
        "ALGA KOMBU ECO 25G",
        "ALGA NORI 25G",
        "GARBANZO ECO",
        "UNKNOWN ITEM",
        "BROKEN PRICE",
        "BROKEN UNIT",
    ]
    assert productos[0].section == "📁 ALGAS"
    assert productos[2].section == "📁 LEGUMBRES"
    assert productos[3].section == "📁 MYSTERY"
    assert productos[4].section == "📁 ALGAS"
    assert productos[0].row_no == 7


def _cfg(fixtures_dir):
    return load_config(fixtures_dir / "lavidaagranel_min.yaml")


def test_map_row_happy_unidades(fixtures_dir):
    cfg = _cfg(fixtures_dir)
    raw = RawRow(kind="product", row_no=7, section="📁 ALGAS",
                 producto="ALGA KOMBU ECO 25G", precio=2.5, unidad="Unidades")
    result = map_row(raw, cfg)
    assert result.skip_reason is None
    row = result.row
    assert row.productor == "La Vida a Granel"
    assert row.nombre == "Alga Kombu Eco 25g"
    assert row.precio_base == "2.50"
    assert row.categoria == "Algas y plantas acuáticas"
    assert row.granel is False
    assert row.pesar is False
    assert row.descripcion == "Se pide por unidades, se paga por kg"
    assert row.productor_id == ""
    assert row.temporada is False


def test_map_row_happy_kg(fixtures_dir):
    cfg = _cfg(fixtures_dir)
    raw = RawRow(kind="product", row_no=10, section="📁 LEGUMBRES",
                 producto="GARBANZO ECO", precio=3.2, unidad="kg")
    result = map_row(raw, cfg)
    assert result.skip_reason is None
    assert result.row.precio_base == "3.20"
    assert result.row.granel is True
    assert result.row.pesar is True
    assert result.row.descripcion == ""


def test_map_row_unmapped_category_skipped(fixtures_dir):
    cfg = _cfg(fixtures_dir)
    raw = RawRow(kind="product", row_no=12, section="📁 MYSTERY",
                 producto="UNKNOWN ITEM", precio=1.0, unidad="kg")
    result = map_row(raw, cfg)
    assert result.skip_reason == "unmapped_category"
    assert result.row is None


def test_map_row_missing_price_skipped(fixtures_dir):
    cfg = _cfg(fixtures_dir)
    raw = RawRow(kind="product", row_no=14, section="📁 ALGAS",
                 producto="BROKEN PRICE", precio=None, unidad="kg")
    result = map_row(raw, cfg)
    assert result.skip_reason == "missing_price"


def test_map_row_missing_unit_skipped(fixtures_dir):
    cfg = _cfg(fixtures_dir)
    raw = RawRow(kind="product", row_no=15, section="📁 ALGAS",
                 producto="BROKEN UNIT", precio=1.0, unidad=None)
    result = map_row(raw, cfg)
    assert result.skip_reason == "missing_unit"


def test_map_row_no_current_section_skipped(fixtures_dir):
    cfg = _cfg(fixtures_dir)
    raw = RawRow(kind="product", row_no=6, section=None,
                 producto="ORPHAN", precio=1.0, unidad="kg")
    result = map_row(raw, cfg)
    assert result.skip_reason == "no_section"


from datetime import date


def test_run_integration(tmp_path, fixtures_dir):
    from scripts.preprocess.lavidaagranel import run

    xlsx = tmp_path / "min.xlsx"
    _build_fixture_xlsx(xlsx)
    out_dir = tmp_path / "out"
    log_dir = tmp_path / "logs"

    summary = run(
        xlsx=xlsx,
        config=fixtures_dir / "lavidaagranel_min.yaml",
        out_dir=out_dir,
        log_dir=log_dir,
        today=date(2026, 5, 16),
    )

    assert summary == {
        "read": 6,
        "emitted": 3,
        "skipped_unmapped": 1,
        "skipped_invalid": 2,
        "skipped_dedup": 0,
        "rejected": 0,
    }

    csv_path = out_dir / "2026-05-16-karakolas-lavidaagranel.csv"
    log_path = log_dir / "2026-05-16-preprocess-lavidaagranel.log"
    assert csv_path.exists()
    assert log_path.exists()

    csv_lines = csv_path.read_text(encoding="utf-8").splitlines()
    assert len(csv_lines) == 1 + 3
    assert "Alga Kombu Eco 25g" in csv_lines[1]
    assert "Garbanzo Eco" in csv_lines[3]

    log_text = log_path.read_text(encoding="utf-8")
    assert "unmapped_category" in log_text
    assert "missing_price" in log_text
    assert "missing_unit" in log_text
    assert "SUMMARY" in log_text


def test_run_dedup(tmp_path, fixtures_dir):
    from scripts.preprocess.lavidaagranel import run

    xlsx = tmp_path / "dup.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pedidos Grupo Consumo"
    for _ in range(4):
        ws.append([])
    ws.append(["PRODUCTO", "PRECIO", "UNIDAD"])
    ws.append(["📁 ALGAS"])
    ws.append(["ALGA KOMBU ECO 25G", 2.5, "Unidades"])
    ws.append(["alga kombu eco 25g", 2.5, "Unidades"])
    wb.save(xlsx)

    out_dir = tmp_path / "out"
    log_dir = tmp_path / "logs"
    summary = run(
        xlsx=xlsx,
        config=fixtures_dir / "lavidaagranel_min.yaml",
        out_dir=out_dir,
        log_dir=log_dir,
        today=date(2026, 5, 16),
    )
    assert summary["emitted"] == 1
    assert summary["skipped_dedup"] == 1


def test_run_idempotent(tmp_path, fixtures_dir):
    from scripts.preprocess.lavidaagranel import run

    xlsx = tmp_path / "min.xlsx"
    _build_fixture_xlsx(xlsx)
    out_a = tmp_path / "a"
    out_b = tmp_path / "b"
    log_a = tmp_path / "la"
    log_b = tmp_path / "lb"
    args = dict(xlsx=xlsx, config=fixtures_dir / "lavidaagranel_min.yaml",
                today=date(2026, 5, 16))
    run(out_dir=out_a, log_dir=log_a, **args)
    run(out_dir=out_b, log_dir=log_b, **args)
    csv_a = (out_a / "2026-05-16-karakolas-lavidaagranel.csv").read_bytes()
    csv_b = (out_b / "2026-05-16-karakolas-lavidaagranel.csv").read_bytes()
    assert csv_a == csv_b
