import argparse
import re
import sys
from dataclasses import dataclass
from datetime import date as _date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterator, Literal

if __name__ == "__main__" and __package__ is None:
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

import openpyxl
import yaml

from scripts.preprocess._common import KarakolasRow, validate, write_csv

_DIGIT_LETTER_TOKEN = re.compile(r"^(\d+)([A-Za-z]+)$")


def normalize(name: str) -> str:
    collapsed = " ".join(name.split())
    titled = collapsed.title()
    out_tokens: list[str] = []
    for token in titled.split(" "):
        m = _DIGIT_LETTER_TOKEN.match(token)
        if m:
            out_tokens.append(f"{m.group(1)}{m.group(2).lower()}")
        else:
            out_tokens.append(token)
    return " ".join(out_tokens)


_REQUIRED_TOP_KEYS = {"productor", "categorias", "unidades", "defaults"}
_OPTIONAL_TOP_KEYS = {"productor_id"}
_REQUIRED_DEFAULTS_KEYS = {"destacado", "temporada", "precio_final", "precio_productor"}


@dataclass(frozen=True)
class Config:
    productor: str
    productor_id: str
    categorias: dict[str, str]
    unidades: dict[str, dict[str, Any]]
    defaults: dict[str, Any]


def load_config(path: Path) -> Config:
    raw = yaml.safe_load(path.read_text(encoding="utf-8"))
    if not isinstance(raw, dict):
        raise ValueError(f"config root must be a mapping, got {type(raw).__name__}")

    keys = set(raw.keys())
    missing = _REQUIRED_TOP_KEYS - keys
    if missing:
        raise ValueError(f"missing required top-level keys: {sorted(missing)}")
    unknown = keys - _REQUIRED_TOP_KEYS - _OPTIONAL_TOP_KEYS
    if unknown:
        raise ValueError(f"unknown top-level keys: {sorted(unknown)}")

    defaults = raw["defaults"] or {}
    missing_defaults = _REQUIRED_DEFAULTS_KEYS - set(defaults.keys())
    if missing_defaults:
        raise ValueError(f"missing required defaults keys: {sorted(missing_defaults)}")

    categorias: dict[str, str] = {}
    for k, v in (raw["categorias"] or {}).items():
        if v is None:
            continue
        if not isinstance(v, dict) or "name" not in v:
            raise ValueError(f"categoria {k!r} must be a mapping with key 'name'")
        categorias[k] = v["name"]

    unidades: dict[str, dict[str, Any]] = {}
    for k, v in (raw["unidades"] or {}).items():
        if not isinstance(v, dict):
            raise ValueError(f"unidad {k!r} must be a mapping")
        required = {"granel", "pesar", "descripcion"}
        if not required.issubset(v.keys()):
            raise ValueError(
                f"unidad {k!r} missing keys: {sorted(required - set(v.keys()))}"
            )
        unidades[k.lower()] = {
            "granel": bool(v["granel"]),
            "pesar": bool(v["pesar"]),
            "descripcion": str(v["descripcion"]),
        }

    return Config(
        productor=str(raw["productor"]),
        productor_id=str(raw.get("productor_id", "")),
        categorias=categorias,
        unidades=unidades,
        defaults=defaults,
    )


@dataclass(frozen=True)
class RawRow:
    kind: Literal["product", "section", "blank"]
    row_no: int
    section: str | None
    producto: str | None
    precio: Any
    unidad: str | None


_SHEET_NAME = "Pedidos Grupo Consumo"
_DATA_START_ROW = 6


def iter_rows(xlsx_path: Path) -> Iterator[RawRow]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[_SHEET_NAME]
    current_section: str | None = None
    row_no = 0
    for raw in ws.iter_rows(values_only=True):
        row_no += 1
        if row_no < _DATA_START_ROW:
            continue
        a = raw[0] if len(raw) > 0 else None
        b = raw[1] if len(raw) > 1 else None
        c = raw[2] if len(raw) > 2 else None
        if a is None and b is None and c is None:
            yield RawRow(kind="blank", row_no=row_no, section=current_section,
                         producto=None, precio=None, unidad=None)
            continue
        if isinstance(a, str) and a.startswith("📁"):
            current_section = a.strip()
            yield RawRow(kind="section", row_no=row_no, section=current_section,
                         producto=None, precio=None, unidad=None)
            continue
        yield RawRow(
            kind="product",
            row_no=row_no,
            section=current_section,
            producto=str(a) if a is not None else None,
            precio=b,
            unidad=str(c) if c is not None else None,
        )


@dataclass(frozen=True)
class MappedResult:
    row: KarakolasRow | None
    skip_reason: str | None


def _format_price(value: Any) -> str | None:
    if value is None or value == "":
        return None
    try:
        d = Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None
    return f"{d:.2f}"


def map_row(raw: RawRow, cfg: Config) -> MappedResult:
    if raw.section is None:
        return MappedResult(None, "no_section")
    if raw.section not in cfg.categorias:
        return MappedResult(None, "unmapped_category")
    if raw.producto is None or not str(raw.producto).strip():
        return MappedResult(None, "missing_producto")
    precio = _format_price(raw.precio)
    if precio is None:
        return MappedResult(None, "missing_price")
    if not raw.unidad or raw.unidad.lower() not in cfg.unidades:
        return MappedResult(None, "missing_unit")
    unidad = cfg.unidades[raw.unidad.lower()]
    d = cfg.defaults
    row = KarakolasRow(
        productor=cfg.productor,
        nombre=normalize(raw.producto),
        precio_base=precio,
        categoria=cfg.categorias[raw.section],
        productor_id=cfg.productor_id,
        descripcion=unidad["descripcion"],
        granel=unidad["granel"],
        pesar=unidad["pesar"],
        destacado=bool(d["destacado"]),
        temporada=bool(d["temporada"]),
        precio_final=str(d["precio_final"]),
        precio_productor=str(d["precio_productor"]),
    )
    return MappedResult(row, None)


_SKIP_REASON_BUCKET = {
    "no_section": "skipped_invalid",
    "unmapped_category": "skipped_unmapped",
    "missing_producto": "skipped_invalid",
    "missing_price": "skipped_invalid",
    "missing_unit": "skipped_invalid",
}


def _log(fh, level: str, row_no: int, category: str | None,
         nombre: str | None, msg: str) -> None:
    fh.write(
        f"{level}\trow={row_no}\tcategory={category or '-'}\t"
        f"nombre={nombre or '-'}\t{msg}\n"
    )


def run(
    *,
    xlsx: Path,
    config: Path,
    out_dir: Path,
    log_dir: Path,
    today: _date | None = None,
) -> dict[str, int]:
    today = today or _date.today()
    cfg = load_config(config)
    out_dir.mkdir(parents=True, exist_ok=True)
    log_dir.mkdir(parents=True, exist_ok=True)
    csv_path = out_dir / f"{today.isoformat()}-karakolas-lavidaagranel.csv"
    log_path = log_dir / f"{today.isoformat()}-preprocess-lavidaagranel.log"

    summary = {
        "read": 0,
        "emitted": 0,
        "skipped_unmapped": 0,
        "skipped_invalid": 0,
        "skipped_dedup": 0,
        "rejected": 0,
    }

    emitted_rows: list[KarakolasRow] = []
    seen: set[tuple[str, str]] = set()

    with log_path.open("w", encoding="utf-8") as logfh:
        for raw in iter_rows(xlsx):
            if raw.kind != "product":
                continue
            summary["read"] += 1
            result = map_row(raw, cfg)
            if result.skip_reason is not None:
                bucket = _SKIP_REASON_BUCKET[result.skip_reason]
                summary[bucket] += 1
                _log(logfh, "WARN", raw.row_no, raw.section, raw.producto,
                     result.skip_reason)
                continue
            row = result.row
            key = (row.productor, row.nombre)
            if key in seen:
                summary["skipped_dedup"] += 1
                _log(logfh, "WARN", raw.row_no, raw.section, row.nombre,
                     "duplicate_nombre")
                continue
            errs = validate(row)
            if errs:
                summary["rejected"] += 1
                _log(logfh, "ERROR", raw.row_no, raw.section, row.nombre,
                     "; ".join(errs))
                continue
            seen.add(key)
            emitted_rows.append(row)
            summary["emitted"] += 1

        write_csv(emitted_rows, csv_path)

        logfh.write("\nSUMMARY\n")
        for k in ("read", "emitted", "skipped_unmapped", "skipped_invalid",
                  "skipped_dedup", "rejected"):
            logfh.write(f"  {k:<18} {summary[k]}\n")

    return summary


def _exit_code(summary: dict[str, int]) -> int:
    bad = (summary["skipped_unmapped"] + summary["skipped_invalid"]
           + summary["skipped_dedup"] + summary["rejected"])
    return 1 if bad else 0


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(prog="preprocess.lavidaagranel")
    p.add_argument("--xlsx", type=Path,
                   default=Path("docs/Pedidos_LaVidaAGranel_original.xlsx"))
    p.add_argument("--config", type=Path,
                   default=Path("scripts/preprocess/lavidaagranel.yaml"))
    p.add_argument("--out-dir", type=Path, default=Path("data/output"))
    p.add_argument("--log-dir", type=Path, default=Path("logs"))
    args = p.parse_args(argv)

    try:
        summary = run(
            xlsx=args.xlsx,
            config=args.config,
            out_dir=args.out_dir,
            log_dir=args.log_dir,
        )
    except (OSError, ValueError, yaml.YAMLError) as exc:
        print(f"FATAL: {exc}", file=sys.stderr)
        return 2

    for k, v in summary.items():
        print(f"{k:<18} {v}")
    return _exit_code(summary)


if __name__ == "__main__":
    raise SystemExit(main())
